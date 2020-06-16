# -*- coding: utf-8 -*-
from os.path import expanduser

from PyQt5.QtWidgets import (QMainWindow, QFileDialog, QApplication, QMessageBox)
from serch_in_word import Ui_MainWindow # импорт нашего сгенерированного файла
import sys
import zipfile, re
import os
import random
from openpyxl import Workbook


class mywindow(QMainWindow):
    def __init__(self):
        super(mywindow, self).__init__()

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton.clicked.connect(self.file_serch)
        self.ui.pushButton_3.clicked.connect(self.skan_word_doc)
        self.a = []


    def file_serch(self):
        try:
            self.f_n = QFileDialog.getExistingDirectory(self,'Выберете корневой каталог для поиска', expanduser("~"))
            print('Выбран каталог - ', self.f_n)
        except Exception:
            print('Ошибка при выборе каталога! ')
        return self.f_n

    def btnclick_dir(self, path):
        self.docx = zipfile.ZipFile(path)
        self.content = self.docx.read('word/document.xml').decode('utf-8')
        self.cleaned = re.sub('<(.|\n)*?>', '', self.content)

        self.regexp = r'(г|снт|п|пос)(\.|\,|\s)(\s|\.|.|\D*|\D*\,|\D*.|\D*\s\-|\w*|\W*)(ул|пер|б-р|просп|пр-кт|пр)(\.|\,|\s)' \
                 r'(\D*\s|\D*\,|\D*.|\D*\s\-)([\d*\s]{1,3})'

        self.addres = re.findall(self.regexp, self.cleaned, re.IGNORECASE)
        return self.addres

    def skan_word_doc(self):
        try:
            obnxodFile(self.f_n)
            wd = Workbook()
            ws = wd.active
            ws.title = 'Лист1'
            row = 1
            ws['A' + str(row)] = 'Город'
            ws['B' + str(row)] = 'Населенный пункт'
            ws['C' + str(row)] = 'ул/пр'
            ws['D' + str(row)] = 'Улица'
            ws['E' + str(row)] = 'номер дома'
            ws['F' + str(row)] = 'Путь к файлу'
            for i in a:
                kek = self.btnclick_dir(i)
                for b in kek:
                    # for item in b[0]:
                    row += 1
                    ws['A' + str(row)] = b[0]
                    # sheet['B' + str(row)] = list_values[1]
                    ws['B' + str(row)] = b[2]
                    ws['C' + str(row)] = b[3]
                    # sheet['E' + str(row)] = list_values[4]
                    ws['D' + str(row)] = b[5]
                    ws['E' + str(row)] = b[6]
                    ws['F' + str(row)] = i
            rend_name = random.choice([10,999])
            strok = str(rend_name)
            wd.save('report'+strok+'.xlsx')
            QMessageBox.question(self, 'Готово!', "Программа завершила свою работу",
                                 QMessageBox.Ok)
            a.clear()
        except Exception:
            QMessageBox.question(self, 'Ошибка!', "Выберете для начала корневой каталог для поиска!",
                                 QMessageBox.Ok)

a = []
def obnxodFile(path_naw_lib, level=1):
    print('Уровень = ', level, 'Содержимое: ', os.listdir(path_naw_lib))
    for i in os.listdir(path_naw_lib):
        if os.path.isdir(path_naw_lib+'/' + i):
            # print('Спускаемся в ',path_naw_lib+'/' + i)
            obnxodFile(path_naw_lib+'/' + i, level+1)
            # print('Возвращаемся в ', path_naw_lib)
        else:
            tochka = i.find('.')
            if i[tochka:] == '.docx':
                # print('Найден WORD документ ! - ', path_naw_lib+'/'+i)
                a.append(path_naw_lib+'/'+i)



app = QApplication([])
application = mywindow()
application.show()

sys.exit(app.exec())